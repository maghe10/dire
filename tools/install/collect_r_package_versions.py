#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import datetime as dt
import platform
import re
import shutil
import subprocess
from pathlib import Path


DEFAULT_PROJECT_DIR = (
    Path.home()
    / "OneDrive - Västra Götalandsregionen"
    / "git"
    / "dire"
    / "R"
)

DEFAULT_OUTPUT_DIR = (
    Path.home()
    / "OneDrive - Västra Götalandsregionen"
    / "DIRE"
    / "Analyser"
    / "processed"
    / "R"
)


def timestamp() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def find_rscript(user_rscript: str | None = None) -> str:
    if user_rscript:
        p = Path(user_rscript)
        if p.exists():
            return str(p)
        raise FileNotFoundError(f"Rscript not found at: {user_rscript}")

    rscript = shutil.which("Rscript")
    if rscript:
        return rscript

    common_paths = [
        r"C:\Users\xhessm\AppData\Local\Programs\R\R-4.5.2\bin\Rscript.exe",
        r"C:\Program Files\R\R-4.5.2\bin\Rscript.exe"
    ]

    for path in common_paths:
        if Path(path).exists():
            return path

    raise FileNotFoundError(
        "Could not find Rscript. Add R to PATH or pass --rscript "
        r'"C:\Program Files\R\R-4.5.2\bin\Rscript.exe"'
    )


def run_rscript(rscript: str, r_code: str, timeout: int = 180) -> str:
    import tempfile

    with tempfile.NamedTemporaryFile(
        mode="w",
        suffix=".R",
        delete=False,
        encoding="utf-8"
    ) as handle:
        script_path = Path(handle.name)
        handle.write(r_code)

    try:
        result = subprocess.run(
            [rscript, "--vanilla", str(script_path)],
            text=True,
            capture_output=True,
            timeout=timeout,
            check=False,
        )

        if result.returncode != 0:
            raise RuntimeError(
                "Rscript failed.\n\n"
                f"Command: {rscript} --vanilla {script_path}\n\n"
                f"Return code: {result.returncode}\n\n"
                f"STDOUT:\n{result.stdout}\n\n"
                f"STDERR:\n{result.stderr}"
            )

        return result.stdout

    finally:
        try:
            script_path.unlink()
        except OSError:
            pass



def scan_used_r_packages(project_dir: Path) -> dict[str, set[str]]:
    package_sources: dict[str, set[str]] = {}

    r_files = []
    for pattern in ("*.R", "*.r", "*.Rmd", "*.rmd", "*.qmd"):
        r_files.extend(project_dir.rglob(pattern))

    library_patterns = [
        re.compile(r"\blibrary\s*\(\s*['\"]?([A-Za-z][A-Za-z0-9._]+)['\"]?\s*\)"),
        re.compile(r"\brequire\s*\(\s*['\"]?([A-Za-z][A-Za-z0-9._]+)['\"]?\s*\)"),
    ]

    namespace_pattern = re.compile(r"\b([A-Za-z][A-Za-z0-9._]+)::")

    for file_path in r_files:
        try:
            text = file_path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            continue

        rel_path = str(file_path.relative_to(project_dir))

        for pattern in library_patterns:
            for match in pattern.finditer(text):
                pkg = match.group(1)
                package_sources.setdefault(pkg, set()).add(rel_path)

        for match in namespace_pattern.finditer(text):
            pkg = match.group(1)
            package_sources.setdefault(pkg, set()).add(rel_path)

    return package_sources


def collect_installed_packages(rscript: str, output_file: Path) -> None:
    r_code = rf'''
ip <- as.data.frame(installed.packages(), stringsAsFactors = FALSE)

out <- data.frame(
  package = ip$Package,
  version = ip$Version,
  lib_path = ip$LibPath,
  priority = ifelse(is.na(ip$Priority), "", ip$Priority),
  depends = ifelse(is.na(ip$Depends), "", ip$Depends),
  imports = ifelse(is.na(ip$Imports), "", ip$Imports),
  linking_to = ifelse(is.na(ip$LinkingTo), "", ip$LinkingTo),
  suggests = ifelse(is.na(ip$Suggests), "", ip$Suggests),
  license = ifelse(is.na(ip$License), "", ip$License),
  built = ifelse(is.na(ip$Built), "", ip$Built),
  stringsAsFactors = FALSE
)

out <- out[order(tolower(out$package), out$lib_path), ]

write.table(
  out,
  file = "{output_file.as_posix()}",
  sep = ";",
  row.names = FALSE,
  col.names = TRUE,
  quote = TRUE,
  na = "",
  fileEncoding = "UTF-8"
)
'''
    run_rscript(rscript, r_code)


def collect_session_info(rscript: str, output_file: Path) -> None:
    r_code = r'''
cat("Timestamp:", format(Sys.time(), "%Y-%m-%d %H:%M:%S %Z"), "\n")
cat("R executable:", R.home("bin"), "\n")
cat("R home:", R.home(), "\n\n")
print(sessionInfo())
'''
    text = run_rscript(rscript, r_code)
    output_file.write_text(text, encoding="utf-8", errors="replace")


def read_semicolon_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        return list(reader)


def write_semicolon_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=fieldnames,
            delimiter=";",
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(rows)


def make_used_package_table(
    package_sources: dict[str, set[str]],
    installed_rows: list[dict[str, str]],
    output_file: Path,
) -> None:
    installed_by_package = {}

    for row in installed_rows:
        pkg = row.get("package", "")
        if pkg and pkg not in installed_by_package:
            installed_by_package[pkg] = row

    rows = []

    for pkg in sorted(package_sources, key=str.lower):
        installed = installed_by_package.get(pkg, {})

        rows.append(
            {
                "package": pkg,
                "version": installed.get("version", ""),
                "status": "installed" if installed else "not found in installed.packages()",
                "lib_path": installed.get("lib_path", ""),
                "built": installed.get("built", ""),
                "source_files": "; ".join(sorted(package_sources[pkg])),
            }
        )

    write_semicolon_csv(
        output_file,
        rows,
        fieldnames=[
            "package",
            "version",
            "status",
            "lib_path",
            "built",
            "source_files",
        ],
    )


def try_write_excel(
    used_csv: Path,
    installed_csv: Path,
    output_xlsx: Path,
) -> None:
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        print("openpyxl not installed; skipping Excel file.")
        print("Install with: conda install openpyxl")
        return

    def add_sheet_from_csv(wb, sheet_name: str, csv_path: Path):
        rows = read_semicolon_csv(csv_path)
        ws = wb.create_sheet(sheet_name)

        if not rows:
            return

        headers = list(rows[0].keys())
        ws.append(headers)

        for row in rows:
            ws.append([row.get(h, "") for h in headers])

        end_col = openpyxl.utils.get_column_letter(len(headers))
        end_row = len(rows) + 1
        table_ref = f"A1:{end_col}{end_row}"

        table = Table(displayName=sheet_name.replace(" ", "_"), ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            ws.column_dimensions[col_letter].width = min(max_length + 2, 80)

    wb = Workbook()
    wb.remove(wb.active)

    add_sheet_from_csv(wb, "Used R packages", used_csv)
    add_sheet_from_csv(wb, "Installed R packages", installed_csv)

    wb.save(output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Collect R package versions from Windows R and write publication-friendly tables."
    )

    parser.add_argument(
        "--project-dir",
        default=str(DEFAULT_PROJECT_DIR),
        help=f"Directory with R scripts to scan. Default: {DEFAULT_PROJECT_DIR}",
    )

    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help=f"Output directory. Default: {DEFAULT_OUTPUT_DIR}",
    )

    parser.add_argument(
        "--rscript",
        default=None,
        help=r'Optional path to Rscript.exe, e.g. "C:\Program Files\R\R-4.5.2\bin\Rscript.exe"',
    )

    args = parser.parse_args()

    project_dir = Path(args.project_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not project_dir.exists():
        raise FileNotFoundError(f"Project directory not found: {project_dir}")

    rscript = find_rscript(args.rscript)

    ts = timestamp()
    host = platform.node() or "unknown_host"

    installed_csv = output_dir / f"r_installed_packages_{host}_{ts}.csv"
    used_csv = output_dir / f"r_used_packages_{host}_{ts}.csv"
    session_txt = output_dir / f"r_session_info_{host}_{ts}.txt"
    excel_file = output_dir / f"r_package_versions_{host}_{ts}.xlsx"

    print(f"Rscript: {rscript}")
    print(f"Scanning R files in: {project_dir}")

    package_sources = scan_used_r_packages(project_dir)

    collect_installed_packages(rscript, installed_csv)
    collect_session_info(rscript, session_txt)

    installed_rows = read_semicolon_csv(installed_csv)
    make_used_package_table(package_sources, installed_rows, used_csv)

    try_write_excel(
        used_csv=used_csv,
        installed_csv=installed_csv,
        output_xlsx=excel_file,
    )

    print("")
    print(f"Wrote: {used_csv}")
    print(f"Wrote: {installed_csv}")
    print(f"Wrote: {session_txt}")

    if excel_file.exists():
        print(f"Wrote: {excel_file}")

    print("")
    print(f"Number of used packages found in project scripts: {len(package_sources)}")


if __name__ == "__main__":
    main()